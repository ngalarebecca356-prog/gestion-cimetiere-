from ninja import NinjaAPI, Schema
from typing import List, Optional
from .models import Caveau, Defunt, Reservation
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from .emails import envoyer_confirmation_reservation, envoyer_validation_reservation

api = NinjaAPI()


# ─── SCHEMAS ───────────────────────────────────────

class CaveauSchema(Schema):
    id: int
    numero: str
    section: str
    bloc: str
    statut: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None

class CaveauCreateSchema(Schema):
    section: str
    bloc: str
    statut: str = 'disponible'
    latitude: Optional[float] = None
    longitude: Optional[float] = None

class DefuntSchema(Schema):
    id: int
    nom: str
    prenom: str
    date_deces: str


class ReservationSchema(Schema):
    id: int
    caveau_id: int
    statut: str


class ReservationCreateSchema(Schema):
    caveau_id: int
    nom_defunt: str
    prenom_defunt: str
    date_deces: str


class PaiementSchema(Schema):
    reservation_id: int
    methode: str
    montant: float
    numero_telephone: str


# ─── CAVEAUX ───────────────────────────────────────

@api.get("/caveaux", response=List[CaveauSchema])
def liste_caveaux(request):
    return Caveau.objects.all()


@api.get("/caveaux/statut/{statut}", response=List[CaveauSchema])
def caveaux_par_statut(request, statut: str):
    return Caveau.objects.filter(statut=statut)


@api.post("/caveaux", response=CaveauSchema)
def creer_caveau(request, data: CaveauCreateSchema):
    from django.contrib.gis.geos import Point
    
    # Numéro automatique
    dernier = Caveau.objects.order_by('-id').first()
    if dernier and dernier.numero.isdigit():
        prochain_num = str(int(dernier.numero) + 1).zfill(3)
    else:
        prochain_num = "001"
    
    caveau = Caveau(
        numero=prochain_num,
        section=data.section,
        bloc=data.bloc,
        statut=data.statut,
    )
    if data.latitude and data.longitude:
        caveau.localisation = Point(data.longitude, data.latitude)
    caveau.save()
    return caveau


@api.put("/caveaux/{caveau_id}/statut")
def modifier_statut_caveau(request, caveau_id: int, statut: str):
    caveau = get_object_or_404(Caveau, id=caveau_id)
    caveau.statut = statut
    caveau.save()
    return {"message": f"Statut mis à jour : {statut}"}


# ─── RESERVATIONS ──────────────────────────────────

@api.get("/reservations", response=List[ReservationSchema])
def liste_reservations(request):
    return Reservation.objects.all()


@api.post("/reservations")
def creer_reservation(request, data: ReservationCreateSchema):
    from datetime import date
    caveau = get_object_or_404(Caveau, id=data.caveau_id)
    if caveau.statut != 'disponible':
        return api.create_response(request, {"error": "Caveau non disponible"}, status=400)
    defunt = Defunt.objects.create(
        nom=data.nom_defunt,
        prenom=data.prenom_defunt,
        date_deces=data.date_deces,
        caveau=caveau
    )
    caveau.statut = 'reserve'
    caveau.save()
    reservation = Reservation.objects.create(
        client=request.user if request.user.is_authenticated else User.objects.first(),
        caveau=caveau,
        defunt=defunt,
        statut='en_attente'
    )
    try:
        envoyer_confirmation_reservation(reservation)
    except:
        pass
    return {"message": "Réservation créée", "id": reservation.id}


@api.put("/reservations/{reservation_id}/valider")
def valider_reservation(request, reservation_id: int):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    reservation.statut = 'validee'
    reservation.caveau.statut = 'occupe'
    reservation.caveau.save()
    reservation.save()
    try:
        envoyer_validation_reservation(reservation)
    except:
        pass
    return {"message": "Réservation validée"}


# ─── PAIEMENT SIMULATION ───────────────────────────

@api.post("/paiement/simuler")
def simuler_paiement(request, data: PaiementSchema):
    methodes_valides = ['airtel_money', 'mpesa', 'especes', 'virement']
    if data.methode not in methodes_valides:
        return {"error": "Méthode de paiement invalide"}
    reservation = get_object_or_404(Reservation, id=data.reservation_id)
    return {
        "status": "success",
        "message": f"Paiement simulé via {data.methode}",
        "montant": data.montant,
        "numero": data.numero_telephone,
        "reservation_id": reservation.id,
        "transaction_id": f"TXN-{reservation.id}-2026"
    }

from django.http import HttpResponse
from .facture import generer_facture


@api.get("/reservations/{reservation_id}/facture")
def telecharger_facture(request, reservation_id: int):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    buffer = generer_facture(reservation)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="facture_{reservation_id}.pdf"'
    return response

import folium
from django.http import HttpResponse as DjangoResponse


@api.get("/carte")
def carte_caveaux(request):
    carte = folium.Map(location=[-4.7761, 11.8635], zoom_start=15)
    
    caveaux = Caveau.objects.all()
    couleurs = {
        'disponible': 'green',
        'reserve': 'orange',
        'occupe': 'red',
        'non_exploitable': 'gray',
    }
    
    for i, c in enumerate(caveaux):
        if c.localisation:
            lat = c.localisation.y
            lng = c.localisation.x
        else:
            lat = -4.7761 + (i * 0.0003)
            lng = 11.8635 + (i * 0.0003)
        
        couleur = couleurs.get(c.statut, 'blue')
        folium.CircleMarker(
            location=[lat, lng],
            radius=12,
            color=couleur,
            fill=True,
            fill_color=couleur,
            fill_opacity=0.8,
            popup=folium.Popup(
                f"<b>Caveau N° {c.numero}</b><br>"
                f"Section {c.section} — Bloc {c.bloc}<br>"
                f"Statut : {c.statut}",
                max_width=200
            )
        ).add_to(carte)

    # Clic sur la carte via folium natif
    carte.add_child(folium.LatLngPopup())

    # Barre d'info en bas
    legende = """
    <div style="
        position: fixed; bottom: 20px; left: 50%;
        transform: translateX(-50%);
        background: #1a1a1a; color: #c9a84c;
        padding: 12px 20px; border-radius: 8px;
        font-family: monospace; font-size: 13px;
        border: 1px solid #c9a84c;
        z-index: 9999;
        box-shadow: 0 4px 20px rgba(0,0,0,0.5);
    ">
        🗺️ Cliquez sur la carte pour voir les coordonnees du point
        &nbsp;&nbsp;|&nbsp;&nbsp;
        🟢 Disponible &nbsp;
        🟠 Reserve &nbsp;
        🔴 Occupe &nbsp;
        ⚫ Non exploitable
    </div>
    """
    carte.get_root().html.add_child(folium.Element(legende))
    
    return DjangoResponse(carte._repr_html_(), content_type='text/html')

from .models import Concession, Exhumation
from datetime import datetime


class ConcessionSchema(Schema):
    id: int
    caveau_id: int
    type_concession: str
    date_debut: str
    date_fin: Optional[str] = None


class ConcessionCreateSchema(Schema):
    caveau_id: int
    type_concession: str
    date_debut: str
    date_fin: Optional[str] = None


class ExhumationSchema(Schema):
    id: int
    caveau_id: int
    defunt_id: int
    motif: str
    statut: str


class ExhumationCreateSchema(Schema):
    caveau_id: int
    defunt_id: int
    motif: str


# ─── CONCESSIONS ───────────────────────────────────

@api.get("/concessions", response=List[ConcessionSchema])
def liste_concessions(request):
    return Concession.objects.all()


@api.post("/concessions")
def creer_concession(request, data: ConcessionCreateSchema):
    caveau = get_object_or_404(Caveau, id=data.caveau_id)
    client = User.objects.first()
    concession = Concession.objects.create(
        caveau=caveau,
        client=client,
        type_concession=data.type_concession,
        date_debut=data.date_debut,
        date_fin=data.date_fin,
    )
    return {"message": "Concession créée", "id": concession.id}


@api.put("/concessions/{concession_id}/renouveler")
def renouveler_concession(request, concession_id: int, nouvelle_date_fin: str):
    concession = get_object_or_404(Concession, id=concession_id)
    concession.date_fin = nouvelle_date_fin
    concession.renouvele = True
    concession.save()
    return {"message": "Concession renouvelée"}


# ─── EXHUMATIONS ───────────────────────────────────

@api.get("/exhumations", response=List[ExhumationSchema])
def liste_exhumations(request):
    return Exhumation.objects.all()


@api.post("/exhumations")
def creer_exhumation(request, data: ExhumationCreateSchema):
    caveau = get_object_or_404(Caveau, id=data.caveau_id)
    defunt = get_object_or_404(Defunt, id=data.defunt_id)
    demandeur = User.objects.first()
    exhumation = Exhumation.objects.create(
        caveau=caveau,
        defunt=defunt,
        demandeur=demandeur,
        motif=data.motif,
    )
    return {"message": "Demande d'exhumation créée", "id": exhumation.id}


@api.put("/exhumations/{exhumation_id}/valider")
def valider_exhumation(request, exhumation_id: int):
    exhumation = get_object_or_404(Exhumation, id=exhumation_id)
    exhumation.statut = 'validee'
    exhumation.date_validation = datetime.now()
    exhumation.save()
    return {"message": "Exhumation validée"}      

@api.get("/dashboard")
def tableau_de_bord(request):
    total_caveaux = Caveau.objects.count()
    disponibles = Caveau.objects.filter(statut='disponible').count()
    reserves = Caveau.objects.filter(statut='reserve').count()
    occupes = Caveau.objects.filter(statut='occupe').count()
    non_exploitables = Caveau.objects.filter(statut='non_exploitable').count()
    
    taux_occupation = (occupes / total_caveaux * 100) if total_caveaux > 0 else 0
    
    total_reservations = Reservation.objects.count()
    reservations_validees = Reservation.objects.filter(statut='validee').count()
    reservations_attente = Reservation.objects.filter(statut='en_attente').count()
    
    total_concessions = Concession.objects.count()
    total_exhumations = Exhumation.objects.count()

    return {
        "caveaux": {
            "total": total_caveaux,
            "disponibles": disponibles,
            "reserves": reserves,
            "occupes": occupes,
            "non_exploitables": non_exploitables,
            "taux_occupation": round(taux_occupation, 2),
        },
        "reservations": {
            "total": total_reservations,
            "validees": reservations_validees,
            "en_attente": reservations_attente,
        },
        "concessions": total_concessions,
        "exhumations": total_exhumations,
        "revenus_estimes": occupes * 150,
    }

import csv
import io
from django.http import HttpResponse as DjangoResponse


@api.get("/export/caveaux/csv")
def export_caveaux_csv(request):
    response = DjangoResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="caveaux.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Numero', 'Section', 'Bloc', 'Statut'])
    
    for c in Caveau.objects.all():
        writer.writerow([c.id, c.numero, c.section, c.bloc, c.statut])
    
    return response


@api.get("/export/reservations/csv")
def export_reservations_csv(request):
    response = DjangoResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reservations.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Caveau', 'Client', 'Statut', 'Date'])
    
    for r in Reservation.objects.all():
        writer.writerow([
            r.id,
            r.caveau.numero,
            r.client.username,
            r.statut,
            r.date_reservation.strftime('%d/%m/%Y')
        ])
    
    return response     

import openpyxl
from openpyxl.styles import Font, PatternFill


@api.get("/export/caveaux/excel")
def export_caveaux_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caveaux"

    # En-têtes
    headers = ['ID', 'Numero', 'Section', 'Bloc', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    # Données
    for row, c in enumerate(Caveau.objects.all(), 2):
        ws.cell(row=row, column=1, value=c.id)
        ws.cell(row=row, column=2, value=c.numero)
        ws.cell(row=row, column=3, value=c.section)
        ws.cell(row=row, column=4, value=c.bloc)
        ws.cell(row=row, column=5, value=c.statut)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = DjangoResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="caveaux.xlsx"'
    return response